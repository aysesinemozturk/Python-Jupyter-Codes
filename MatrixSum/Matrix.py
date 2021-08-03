from __future__ import print_function
from array import *
# import openpyxl module
import openpyxl
from openpyxl import Workbook

class Matrix:
    
    #numberOfRows = 0
    #numberOfColumns = 0
    #values = [5][5]p'p

    # A constructor gets again matrix's attributes as parameters, however this time it also takes vectorvalues which was
    # given already in the main method is prepared for the matrix's values. 
    # If vectorvalues length is same as our matrix's dimension, it will take it as values.
    # Otherwise, values of the matrix remains as zero, null.

    def __init__(self, title, NumberOfRows, NumberOfColumns, VectorValues = None):
        self.numberOfRows = NumberOfRows
        self.numberOfColumns = NumberOfColumns
        self.title = title

        if(VectorValues == None):
            self.values = [[0 for y in range(self.numberOfColumns)] for x in range(self.numberOfRows)]
        
        elif(len(VectorValues) != NumberOfColumns*NumberOfRows):
            print('the vectorValues array is not suitable for the matrix dimension.')
            
            self.values = [[0 for y in range(self.numberOfColumns)] for x in range(self.numberOfRows)]

        else:
            for i in range(self.numberOfRows):
                for j in range(self.numberOfColumns):
                    self.values[i][j] = VectorValues[i*self.numberOfColumns +j]
    

    #sum up two matrixes which are defined in main
    def sum(self, matrix):

        #if size of two matrixes are equal to each other, summation is possible.
        if(self.numberOfRows == matrix.numberOfRows and self.numberOfColumns == matrix.numberOfColumns):

            resultMatrix = Matrix("Result Matrix",self.numberOfRows,self.numberOfColumns)

            for i in range(self.numberOfRows):
                for j in range(self.numberOfColumns):
                    resultMatrix[i][j] = self.values[i][j] + matrix.values[i][j]
            
            return resultMatrix
        else:
            return None
        
    #report result matrix
    def report(self):

        for i in range(self.numberOfRows):
            print('|', end='')

            for j in range(self.numberOfColumns):
                print(self.values[i][j], end='')
                print(',', end='')
            
            print('|')

    #if we want to make our report method accessible from only in this class, we should use __
    def __PrivateReport(self):

            for i in range(self.numberOfRows):
                print('|', end='')

                for j in range(self.numberOfColumns):
                    print(self.values[i][j], end='')
                    print(',', end='')
                
                print('|')
    #However, we are not going to use this function this time.

    # we should use statis method in order to use following method without creating object
    @staticmethod
    def createMatrixUsingExcelFiles(ExcelFileName, title):

        #access existing workbook which means existing excel files
        wb_obj = openpyxl.load_workbook(filename = ExcelFileName, data_only = True)
        #it is for activating workbook
        ws = wb_obj.active

        #now it is time to assign data values in excel files to empty array, then temp matrix
        values = []
        for i in range (ws.max_column):
            for j in range(ws.max_row):
                values.append(ws.cell(j+1,i+1).value) #ws.cell starts with 1 and range starts with zero/ 
                                                      #ws.cell().value gives us access to cell's value.

        tempMatrix = Matrix(title, ws.max_row, ws.max_column, values)
        return tempMatrix

    def saveMatrixInExcelFile(self):

        #create workbook, excel file
        wb = Workbook()
        #activate workbook
        ws = wb.active
        ws.title = "Matrix " + self.title

        for i in range(self.numberOfRows):
            for j in range(self.numberOfColumns):
                
                ws.cell(i+1,j+1).value = self.values[i][j]
        
        wb.save(filename = self.title + '.xlsx') #save our workbook



